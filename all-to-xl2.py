import mysql.connector
from openpyxl import Workbook
import argparse

def sanitize_for_excel(value):
    if not isinstance(value, str):
        return value
    # Remove common illegal characters for Excel
    illegal_chars = set(chr(i) for i in range(32))
    illegal_chars.discard('\t')
    illegal_chars.discard('\n')
    illegal_chars.discard('\r')
    return ''.join(c for c in value if c not in illegal_chars)

def add_sheet_to_workbook(wb, sheet_name, results, column_names):
    # Create a new worksheet with the specified name
    ws = wb.create_sheet(title=sheet_name)

    # Add headers to the worksheet
    for col_num, column_title in enumerate(column_names, 1):
        col_letter = ws.cell(row=1, column=col_num)
        col_letter.value = sanitize_for_excel(column_title)

    # Add data rows to the worksheet
    for row_num, row_data in enumerate(results, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = sanitize_for_excel(cell_data)

def main(output_file):
    # Database connection details
    db_config = {
        'host': 'host1',
        'database': 'db1',
        'user': 'user1',
        'password': 'pass1'
    }

    # Connect to the database
    connection = mysql.connector.connect(**db_config)
    cursor = connection.cursor()

    # List of your multiline complex queries
    queries = [
        """
        -- Your first query here
        """,
        """
        -- Your second query here
        """,
        """
        -- Your third query here
        """,
        """
        -- Your fourth query here
        """,
        """
        -- Your fifth query here
        """
    ]

    # Create a new workbook
    wb = Workbook()
    # Remove the default sheet created by openpyxl
    wb.remove(wb.active)

    # Execute each query and add its results to a new sheet in the workbook
    for index, query in enumerate(queries, 1):
        cursor.execute(query)
        results = cursor.fetchall()
        column_names = [desc[0] for desc in cursor.description]
        add_sheet_to_workbook(wb, f"Query_{index}", results, column_names)

    # Save the workbook to the specified file
    wb.save(output_file)

    # Close the cursor and connection
    cursor.close()
    connection.close()

    print(f"Data saved to {output_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Save multiple database query results to an XLSX file.")
    parser.add_argument("output_file", help="Name of the output XLSX file.")
    args = parser.parse_args()
    
    main(args.output_file)
