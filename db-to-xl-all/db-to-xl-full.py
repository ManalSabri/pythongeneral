import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
import argparse

def sanitize_for_excel(value):
    if not isinstance(value, str):
        return value
    # Remove common illegal characters for Excel
    illegal_chars = set(chr(i) for i in range(32))
    illegal_chars.discard('\t')
    illegal_chars.discard('\n')
    illegal_chars.discard('\r')
    return ''.join(c for c in value if c not in illegal_chars)

def add_sheet_to_workbook(wb, sheet_name, results, column_names):
    # Create a new worksheet with the specified name
    ws = wb.create_sheet(title=sheet_name)

    # Define border and alignment styles
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(wrap_text=True)

    # Add headers to the worksheet with styles
    for col_num, column_title in enumerate(column_names, 1):
        cell = ws.cell(row=1, column=col_num, value=sanitize_for_excel(column_title))
        cell.border = border
        cell.alignment = alignment

    # Add data rows to the worksheet with styles
    for row_num, row_data in enumerate(results, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=sanitize_for_excel(cell_data))
            cell.border = border
            cell.alignment = alignment

    # Set default column width
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 60:
            adjusted_width = 60
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

def read_queries_from_file(filename):
    with open(filename, 'r') as file:
        content = file.read()
        sections = content.split(">>>>")[1:]  # Split by ">>>>" and ignore the first empty section
        queries = [(section.split('\n', 1)[0], section.split('\n', 1)[1].strip()) for section in sections]
    return queries

def main(input_file, output_file):
    # Database connection details
    db_config = {
        'host': 'host1',
        'database': 'db1',
        'user': 'user1',
        'password': 'pass1'
    }

    # Connect to the database
    connection = mysql.connector.connect(**db_config)
    cursor = connection.cursor()

    # Read queries and their titles from the input file
    queries = read_queries_from_file(input_file)

    # Create a new workbook
    wb = Workbook()
    # Remove the default sheet created by openpyxl
    wb.remove(wb.active)

    # Execute each query and add its results to a new sheet in the workbook
    for title, query in queries:
        cursor.execute(query)
        results = cursor.fetchall()
        column_names = [desc[0] for desc in cursor.description]
        add_sheet_to_workbook(wb, title, results, column_names)

    # Save the workbook to the specified file
    wb.save(output_file)

    # Close the cursor and connection
    cursor.close()
    connection.close()

    print(f"Data saved to {output_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Save multiple database query results to an XLSX file.")
    parser.add_argument("input_file", help="Name of the input text file containing queries and titles.")
    parser.add_argument("output_file", help="Name of the output XLSX file.")
    args = parser.parse_args()
    
    main(args.input_file, args.output_file)
